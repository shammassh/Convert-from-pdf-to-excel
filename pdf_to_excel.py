import os
from pathlib import Path
import openpyxl
from openpyxl import Workbook
import pdfplumber
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import threading
import json
from collections import OrderedDict

def get_pdf_files(folder_path):
    """
    Get all PDF files from the specified folder.
    
    Args:
        folder_path: Path to the folder containing PDF files
        
    Returns:
        List of PDF file paths
    """
    pdf_files = []
    
    if not os.path.exists(folder_path):
        print(f"Error: Folder '{folder_path}' does not exist.")
        return pdf_files
    
    for file in os.listdir(folder_path):
        if file.lower().endswith('.pdf'):
            pdf_files.append(os.path.join(folder_path, file))
    
    return sorted(pdf_files)

def extract_all_fields_from_pdf(pdf_path, max_pages=3):
    """
    Extract all possible fields from a PDF file.
    
    Args:
        pdf_path: Full path to the PDF file
        max_pages: Maximum number of pages to analyze (default: 3)
        
    Returns:
        Dictionary of field_name: value pairs
    """
    fields = OrderedDict()
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Limit analysis to first few pages for performance
            pages_to_analyze = min(max_pages, len(pdf.pages))
            
            for page_num in range(pages_to_analyze):
                page = pdf.pages[page_num]
                
                # Extract text
                text = page.extract_text() or ""
                
                # Extract tables
                tables = page.extract_tables()
                
                # Method 1: Extract from tables (column headers and first data row)
                if tables:
                    for table in tables:
                        if not table or len(table) < 2:
                            continue
                        
                        # Get headers (first row)
                        headers = table[0] if table else []
                        
                        # Try to pair headers with values from subsequent rows
                        for col_idx, header in enumerate(headers):
                            if not header or not str(header).strip():
                                continue
                            
                            header_clean = str(header).strip()
                            
                            # Look for values in the same column
                            for row_idx in range(1, min(6, len(table))):  # Check first 5 data rows
                                if len(table[row_idx]) > col_idx:
                                    value = table[row_idx][col_idx]
                                    if value and str(value).strip():
                                        field_key = f"{header_clean}"
                                        if field_key not in fields:
                                            fields[field_key] = str(value).strip()
                                        break
                
                # Method 2: Extract key-value pairs from text
                # Common patterns: "Label: Value", "Label Value", "Label\nValue"
                lines = text.split('\n')
                
                for i, line in enumerate(lines):
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Pattern: "Label: Value"
                    if ':' in line:
                        parts = line.split(':', 1)
                        if len(parts) == 2:
                            key = parts[0].strip()
                            value = parts[1].strip()
                            if key and value and len(key) < 50 and len(value) < 200:
                                if key not in fields:
                                    fields[key] = value
                    
                    # Pattern: Look for common invoice fields
                    common_patterns = [
                        (r'Invoice\s*(?:Number|#|No\.?)\s*[:\s]\s*(.+)', 'Invoice Number'),
                        (r'Invoice\s*Date\s*[:\s]\s*(.+)', 'Invoice Date'),
                        (r'Due\s*Date\s*[:\s]\s*(.+)', 'Due Date'),
                        (r'Total\s*Amount\s*[:\s]?\s*(?:USD|usd|\$)?\s*([0-9,]+\.?[0-9]*)', 'Total Amount'),
                        (r'Subtotal\s*[:\s]?\s*(?:USD|usd|\$)?\s*([0-9,]+\.?[0-9]*)', 'Subtotal'),
                        (r'Tax\s*[:\s]?\s*(?:USD|usd|\$)?\s*([0-9,]+\.?[0-9]*)', 'Tax'),
                        (r'Amount\s*Due\s*[:\s]?\s*(?:USD|usd|\$)?\s*([0-9,]+\.?[0-9]*)', 'Amount Due'),
                        (r'Customer\s*(?:Name|ID)?\s*[:\s]\s*(.+)', 'Customer'),
                        (r'Vendor\s*(?:Name|ID)?\s*[:\s]\s*(.+)', 'Vendor'),
                        (r'PO\s*(?:Number|#)?\s*[:\s]\s*(.+)', 'PO Number'),
                    ]
                    
                    for pattern, field_name in common_patterns:
                        match = re.search(pattern, line, re.IGNORECASE)
                        if match and field_name not in fields:
                            value = match.group(1).strip()
                            if value:
                                fields[field_name] = value
                
                # Method 3: Look for standalone labeled fields across multiple lines
                for i, line in enumerate(lines):
                    # Check if this line looks like a label (short, ends with specific keywords)
                    if len(line) < 50 and any(keyword in line.lower() for keyword in 
                        ['number', 'date', 'name', 'id', 'code', 'amount', 'total', 'address', 'email', 'phone']):
                        # Check next line for potential value
                        if i + 1 < len(lines):
                            next_line = lines[i + 1].strip()
                            if next_line and len(next_line) < 200:
                                label = line.strip().rstrip(':')
                                if label and label not in fields:
                                    fields[label] = next_line
                
    except Exception as e:
        print(f"Error extracting fields from {os.path.basename(pdf_path)}: {str(e)}")
    
    return fields

def extract_field_from_pdf(pdf_path, field_patterns):
    """
    Extract specific field(s) from PDF based on field patterns.
    
    Args:
        pdf_path: Full path to the PDF file
        field_patterns: List of field names/patterns to search for
        
    Returns:
        Dictionary of found field values
    """
    results = {}
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:3]:  # Check first 3 pages
                text = page.extract_text() or ""
                tables = page.extract_tables()
                
                # Search in tables
                if tables:
                    for table in tables:
                        if not table:
                            continue
                        
                        for row in table:
                            if not row:
                                continue
                            
                            for cell in row:
                                if not cell:
                                    continue
                                
                                cell_str = str(cell)
                                for pattern in field_patterns:
                                    if pattern.lower() in cell_str.lower() and pattern not in results:
                                        # Try to find value in same row or next rows
                                        row_idx = table.index(row)
                                        col_idx = row.index(cell)
                                        
                                        # Look in same column, next rows
                                        for data_row in table[row_idx + 1:row_idx + 5]:
                                            if len(data_row) > col_idx and data_row[col_idx]:
                                                value = str(data_row[col_idx]).strip()
                                                if value and value != cell_str:
                                                    results[pattern] = value
                                                    break
                
                # Search in text
                for pattern in field_patterns:
                    if pattern in results:
                        continue
                    
                    # Try various regex patterns
                    regex_patterns = [
                        rf'{re.escape(pattern)}\s*[:\s]\s*(.+?)(?:\n|$)',
                        rf'{re.escape(pattern)}.*?([0-9,]+\.?[0-9]*)',
                    ]
                    
                    for regex in regex_patterns:
                        match = re.search(regex, text, re.IGNORECASE | re.DOTALL)
                        if match:
                            value = match.group(1).strip()
                            if value:
                                results[pattern] = value
                                break
                
                # If we found all patterns, break early
                if len(results) == len(field_patterns):
                    break
    
    except Exception as e:
        print(f"Error extracting field from {os.path.basename(pdf_path)}: {str(e)}")
    
    return results

def extract_total_amount(pdf_path):
    """
    Extract the total amount from a PDF file by looking for 'Total Amount' column.
    
    Args:
        pdf_path: Full path to the PDF file
        
    Returns:
        Total amount as string or 'N/A' if not found
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Try to extract tables from all pages
            for page in pdf.pages:
                tables = page.extract_tables()
                
                # Check if tables exist
                if tables:
                    for table in tables:
                        if not table:
                            continue
                        
                        # Look for 'Total Amount' in the table
                        for row_idx, row in enumerate(table):
                            if not row:
                                continue
                            
                            # Check each cell for 'Total Amount'
                            for col_idx, cell in enumerate(row):
                                if cell and 'Total Amount' in str(cell):
                                    # Try to find the value in the same column
                                    for data_row in table[row_idx + 1:]:
                                        if data_row and len(data_row) > col_idx:
                                            value = data_row[col_idx]
                                            if value and str(value).strip():
                                                # Clean and return the value
                                                cleaned = str(value).replace(',', '').replace('USD', '').strip()
                                                if re.match(r'^[0-9.]+$', cleaned):
                                                    return cleaned
                
                # Extract text and look for the pattern
                text = page.extract_text() or ""
                
                # Look for "Total Amount" followed by optional due date and USD amount
                # Pattern: Total Amount ... Due on ... USD 239.40
                pattern = r'Total Amount.*?USD\s*([0-9,]+\.?[0-9]*)'
                match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
                if match:
                    amount = match.group(1).replace(',', '')
                    return amount
                
                # Alternative: Look for USD followed by amount near Total Amount
                lines = text.split('\n')
                for i, line in enumerate(lines):
                    if 'Total Amount' in line:
                        # Check next few lines for USD amount
                        for j in range(i, min(i + 5, len(lines))):
                            usd_match = re.search(r'USD\s*([0-9,]+\.?[0-9]*)', lines[j])
                            if usd_match:
                                amount = usd_match.group(1).replace(',', '')
                                return amount
                
                # Fallback patterns
                patterns = [
                    r'Total Amount[:\s]+USD\s*([0-9,]+\.?[0-9]*)',
                    r'Total Amount[:\s]+([0-9,]+\.?[0-9]*)',
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, text, re.IGNORECASE)
                    if match:
                        amount = match.group(1).replace(',', '')
                        return amount
            
            return 'N/A'
            
    except Exception as e:
        print(f"   ⚠ Error reading {os.path.basename(pdf_path)}: {str(e)}")
        return 'Error'

def write_to_excel(pdf_data, excel_path):
    """
    Write PDF filenames, total amounts, and hyperlinks to an Excel file.
    Prevents duplicate entries by checking existing data.
    
    Args:
        pdf_data: List of tuples (filename, total_amount, full_path)
        excel_path: Path where the Excel file will be saved
    """
    try:
        # Check if Excel file already exists
        existing_files = set()
        if os.path.exists(excel_path):
            try:
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
                
                # Collect existing filenames by iterating only through cells with values
                # This completely ignores empty/deleted rows
                for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
                    cell_value = row[0].value
                    if cell_value and str(cell_value).strip():
                        existing_files.add(cell_value)
                
            except PermissionError:
                print(f"\n❌ ERROR: Cannot open '{excel_path}'")
                print("   The file is currently open in another program.")
                print("   Please close the file and try again.\n")
                return False
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "PDF Files"
            # Add headers
            ws['A1'] = "PDF Filename"
            ws['B1'] = "Total Amount"
            ws['C1'] = "Path to Invoice"
            ws['A1'].font = openpyxl.styles.Font(bold=True)
            ws['B1'].font = openpyxl.styles.Font(bold=True)
            ws['C1'].font = openpyxl.styles.Font(bold=True)
        
        # Filter out duplicates
        new_data = [(name, amount, path) for name, amount, path in pdf_data if name not in existing_files]
        duplicates_count = len(pdf_data) - len(new_data)
        
        if duplicates_count > 0:
            print(f"\n⚠ Skipped {duplicates_count} duplicate file(s)")
        
        if not new_data:
            print("\n⚠ No new files to add (all files already exist in Excel)")
            return True
        
        # Find the next empty row
        start_row = ws.max_row + 1 if ws.max_row > 1 else 2
        
        # Write only new PDF data
        for idx, (pdf_name, total_amount, pdf_path) in enumerate(new_data, start=start_row):
            ws[f'A{idx}'] = pdf_name
            ws[f'B{idx}'] = total_amount
            # Use relative path from Excel file location
            relative_path = os.path.relpath(pdf_path, os.path.dirname(excel_path))
            ws[f'C{idx}'].hyperlink = relative_path
            ws[f'C{idx}'].value = "Open Invoice"
            ws[f'C{idx}'].font = openpyxl.styles.Font(color="0563C1", underline="single")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
        # Save the workbook
        try:
            wb.save(excel_path)
            print(f"\n✓ Successfully wrote {len(new_data)} PDF file(s) to {excel_path}")
            if duplicates_count > 0:
                print(f"  ({duplicates_count} duplicate(s) skipped)")
            return True
        except PermissionError:
            print(f"\n❌ ERROR: Cannot save to '{excel_path}'")
            print("   The file is currently open in another program (Excel, etc.)")
            print("   Please close the file and try again.\n")
            
            # Offer to save with a different name
            retry = input("Would you like to save with a different filename? (yes/no): ").strip().lower()
            if retry in ['yes', 'y']:
                new_path = input("Enter new Excel file path: ").strip()
                if not new_path.lower().endswith('.xlsx'):
                    new_path += '.xlsx'
                return write_to_excel(pdf_data, new_path)
            return False
            
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        return False

def main():
    """
    Main function to run the PDF to Excel application with GUI.
    """
    app = PDFtoExcelApp()
    app.mainloop()

class FieldMappingDialog(tk.Toplevel):
    """Dialog for selecting and mapping PDF fields to Excel columns"""
    
    def __init__(self, parent, sample_fields, existing_mapping=None):
        super().__init__(parent)
        
        self.title("Field Mapping Configuration")
        self.geometry("900x650")
        self.resizable(True, True)
        
        self.sample_fields = sample_fields
        self.field_mapping = existing_mapping if existing_mapping else []
        self.result = None  # Will store the selected field mapping
        
        # Make dialog modal
        self.transient(parent)
        self.grab_set()
        
        self.create_widgets()
        
        # Center on parent
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")
        
    def create_widgets(self):
        # Header
        header_frame = tk.Frame(self, bg="#2c3e50", height=60)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame,
            text="Configure Field Mapping",
            font=("Arial", 18, "bold"),
            bg="#2c3e50",
            fg="white"
        )
        title_label.pack(pady=15)
        
        # Instructions
        instruction_frame = tk.Frame(self, bg="#ecf0f1", relief="solid", bd=1)
        instruction_frame.pack(fill="x", padx=20, pady=(20, 10))
        
        instruction_text = (
            "Select which fields to extract from PDFs and map to Excel columns.\n"
            "Preview shows fields detected from a sample PDF. Add the fields you want to track."
        )
        instruction_label = tk.Label(
            instruction_frame,
            text=instruction_text,
            font=("Arial", 10),
            bg="#ecf0f1",
            fg="#2c3e50",
            justify="left",
            padx=10,
            pady=10
        )
        instruction_label.pack()
        
        # Main content with two panels
        content_frame = tk.Frame(self, bg="#f0f0f0")
        content_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Left panel - Available fields
        left_panel = tk.LabelFrame(
            content_frame,
            text="Available Fields (from sample PDF)",
            font=("Arial", 11, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50",
            padx=10,
            pady=10
        )
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Search box
        search_frame = tk.Frame(left_panel, bg="#f0f0f0")
        search_frame.pack(fill="x", pady=(0, 10))
        
        tk.Label(search_frame, text="Search:", font=("Arial", 9), bg="#f0f0f0").pack(side="left", padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.filter_fields)
        search_entry = tk.Entry(search_frame, textvariable=self.search_var, font=("Arial", 9))
        search_entry.pack(side="left", fill="x", expand=True)
        
        # Listbox with scrollbar
        list_frame = tk.Frame(left_panel)
        list_frame.pack(fill="both", expand=True)
        
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")
        
        self.available_listbox = tk.Listbox(
            list_frame,
            font=("Consolas", 9),
            yscrollcommand=scrollbar.set,
            selectmode=tk.SINGLE,
            bg="white"
        )
        self.available_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.available_listbox.yview)
        
        # Populate available fields
        self.all_fields = list(self.sample_fields.keys())
        self.populate_available_fields()
        
        # Field value preview
        preview_frame = tk.Frame(left_panel, bg="#f0f0f0")
        preview_frame.pack(fill="x", pady=(10, 0))
        
        tk.Label(preview_frame, text="Sample Value:", font=("Arial", 9, "bold"), bg="#f0f0f0").pack(anchor="w")
        self.preview_label = tk.Label(
            preview_frame,
            text="Select a field to see its value",
            font=("Arial", 9),
            bg="white",
            fg="#555",
            relief="sunken",
            padx=5,
            pady=5,
            wraplength=350,
            justify="left"
        )
        self.preview_label.pack(fill="x", pady=(5, 0))
        
        self.available_listbox.bind('<<ListboxSelect>>', self.show_field_preview)
        self.available_listbox.bind('<Double-Button-1>', self.add_field)
        
        # Middle panel - Action buttons
        button_panel = tk.Frame(content_frame, bg="#f0f0f0", width=80)
        button_panel.pack(side="left", fill="y", padx=10)
        button_panel.pack_propagate(False)
        
        # Spacer
        tk.Frame(button_panel, bg="#f0f0f0", height=80).pack()
        
        add_btn = tk.Button(
            button_panel,
            text="Add →",
            command=self.add_field,
            font=("Arial", 10, "bold"),
            bg="#27ae60",
            fg="white",
            cursor="hand2",
            width=10,
            pady=8
        )
        add_btn.pack(pady=5)
        
        remove_btn = tk.Button(
            button_panel,
            text="← Remove",
            command=self.remove_field,
            font=("Arial", 10, "bold"),
            bg="#e74c3c",
            fg="white",
            cursor="hand2",
            width=10,
            pady=8
        )
        remove_btn.pack(pady=5)
        
        tk.Frame(button_panel, bg="#f0f0f0", height=20).pack()
        
        move_up_btn = tk.Button(
            button_panel,
            text="Move Up ↑",
            command=self.move_up,
            font=("Arial", 9),
            bg="#3498db",
            fg="white",
            cursor="hand2",
            width=10,
            pady=5
        )
        move_up_btn.pack(pady=5)
        
        move_down_btn = tk.Button(
            button_panel,
            text="Move Down ↓",
            command=self.move_down,
            font=("Arial", 9),
            bg="#3498db",
            fg="white",
            cursor="hand2",
            width=10,
            pady=5
        )
        move_down_btn.pack(pady=5)
        
        # Right panel - Selected fields (will become Excel columns)
        right_panel = tk.LabelFrame(
            content_frame,
            text="Selected Fields (Excel Columns)",
            font=("Arial", 11, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50",
            padx=10,
            pady=10
        )
        right_panel.pack(side="left", fill="both", expand=True, padx=(10, 0))
        
        # Listbox with scrollbar
        list_frame2 = tk.Frame(right_panel)
        list_frame2.pack(fill="both", expand=True)
        
        scrollbar2 = tk.Scrollbar(list_frame2)
        scrollbar2.pack(side="right", fill="y")
        
        self.selected_listbox = tk.Listbox(
            list_frame2,
            font=("Consolas", 9),
            yscrollcommand=scrollbar2.set,
            selectmode=tk.SINGLE,
            bg="white"
        )
        self.selected_listbox.pack(side="left", fill="both", expand=True)
        scrollbar2.config(command=self.selected_listbox.yview)
        
        # Populate existing mapping
        if self.field_mapping:
            for field in self.field_mapping:
                self.selected_listbox.insert(tk.END, field)
        
        # Column count info
        count_label = tk.Label(
            right_panel,
            text=f"Columns: {len(self.field_mapping)}",
            font=("Arial", 9),
            bg="#f0f0f0",
            fg="#555"
        )
        count_label.pack(pady=(10, 0))
        self.count_label = count_label
        
        # Bottom buttons
        bottom_frame = tk.Frame(self, bg="#f0f0f0")
        bottom_frame.pack(fill="x", padx=20, pady=20)
        
        cancel_btn = tk.Button(
            bottom_frame,
            text="Cancel",
            command=self.cancel,
            font=("Arial", 11),
            bg="#95a5a6",
            fg="white",
            cursor="hand2",
            padx=20,
            pady=8
        )
        cancel_btn.pack(side="right", padx=(10, 0))
        
        save_btn = tk.Button(
            bottom_frame,
            text="Save Mapping",
            command=self.save_mapping,
            font=("Arial", 11, "bold"),
            bg="#27ae60",
            fg="white",
            cursor="hand2",
            padx=20,
            pady=8
        )
        save_btn.pack(side="right")
        
    def populate_available_fields(self):
        """Populate the available fields listbox"""
        self.available_listbox.delete(0, tk.END)
        search_term = self.search_var.get().lower()
        
        for field in self.all_fields:
            if search_term in field.lower():
                self.available_listbox.insert(tk.END, field)
    
    def filter_fields(self, *args):
        """Filter available fields based on search"""
        self.populate_available_fields()
    
    def show_field_preview(self, event):
        """Show preview of selected field value"""
        selection = self.available_listbox.curselection()
        if selection:
            field_name = self.available_listbox.get(selection[0])
            value = self.sample_fields.get(field_name, "N/A")
            self.preview_label.config(text=f"{value}")
    
    def add_field(self, event=None):
        """Add selected field to mapping"""
        selection = self.available_listbox.curselection()
        if selection:
            field_name = self.available_listbox.get(selection[0])
            if field_name not in self.field_mapping:
                self.field_mapping.append(field_name)
                self.selected_listbox.insert(tk.END, field_name)
                self.update_count()
    
    def remove_field(self):
        """Remove selected field from mapping"""
        selection = self.selected_listbox.curselection()
        if selection:
            idx = selection[0]
            field_name = self.selected_listbox.get(idx)
            self.field_mapping.remove(field_name)
            self.selected_listbox.delete(idx)
            self.update_count()
    
    def move_up(self):
        """Move selected field up in order"""
        selection = self.selected_listbox.curselection()
        if selection and selection[0] > 0:
            idx = selection[0]
            # Swap in mapping list
            self.field_mapping[idx], self.field_mapping[idx - 1] = \
                self.field_mapping[idx - 1], self.field_mapping[idx]
            # Refresh listbox
            self.refresh_selected_list()
            self.selected_listbox.selection_set(idx - 1)
    
    def move_down(self):
        """Move selected field down in order"""
        selection = self.selected_listbox.curselection()
        if selection and selection[0] < len(self.field_mapping) - 1:
            idx = selection[0]
            # Swap in mapping list
            self.field_mapping[idx], self.field_mapping[idx + 1] = \
                self.field_mapping[idx + 1], self.field_mapping[idx]
            # Refresh listbox
            self.refresh_selected_list()
            self.selected_listbox.selection_set(idx + 1)
    
    def refresh_selected_list(self):
        """Refresh the selected fields listbox"""
        self.selected_listbox.delete(0, tk.END)
        for field in self.field_mapping:
            self.selected_listbox.insert(tk.END, field)
    
    def update_count(self):
        """Update the column count label"""
        self.count_label.config(text=f"Columns: {len(self.field_mapping)}")
    
    def save_mapping(self):
        """Save the field mapping and close dialog"""
        if not self.field_mapping:
            messagebox.showwarning("Warning", "Please select at least one field!")
            return
        
        self.result = self.field_mapping
        self.destroy()
    
    def cancel(self):
        """Cancel and close dialog"""
        self.result = None
        self.destroy()
    
    def get_result(self):
        """Return the field mapping result"""
        return self.result

class PDFtoExcelApp(tk.Tk):
    def __init__(self):
        super().__init__()
        
        self.title("PDF to Excel Converter")
        self.geometry("700x750")
        self.resizable(False, False)
        
        # Configure style
        self.configure(bg="#f0f0f0")
        
        # Variables
        self.folder_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.available_sheets = []
        self.field_mapping = []  # List of field names to extract
        self.mapping_file = "field_mapping.json"  # File to save mapping
        
        # Load saved mapping if exists
        self.load_field_mapping()
        
        # Create GUI elements
        self.create_widgets()
        
    def create_widgets(self):
        # Header
        header_frame = tk.Frame(self, bg="#2c3e50", height=80)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame, 
            text="PDF to Excel Converter",
            font=("Arial", 24, "bold"),
            bg="#2c3e50",
            fg="white"
        )
        title_label.pack(pady=20)
        
        # Main content frame
        content_frame = tk.Frame(self, bg="#f0f0f0")
        content_frame.pack(fill="both", expand=True, padx=30, pady=30)
        
        # Folder selection
        folder_frame = tk.LabelFrame(
            content_frame,
            text="Select PDF Folder",
            font=("Arial", 12, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50",
            padx=15,
            pady=15
        )
        folder_frame.pack(fill="x", pady=(0, 20))
        
        folder_entry = tk.Entry(
            folder_frame,
            textvariable=self.folder_path,
            font=("Arial", 10),
            width=50
        )
        folder_entry.pack(side="left", padx=(0, 10))
        
        folder_btn = tk.Button(
            folder_frame,
            text="Browse Folder",
            command=self.browse_folder,
            font=("Arial", 10),
            bg="#3498db",
            fg="white",
            cursor="hand2",
            padx=15,
            pady=5
        )
        folder_btn.pack(side="left")
        
        # Field Mapping section
        mapping_frame = tk.LabelFrame(
            content_frame,
            text="Field Mapping Configuration",
            font=("Arial", 12, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50",
            padx=15,
            pady=15
        )
        mapping_frame.pack(fill="x", pady=(0, 20))
        
        mapping_info_frame = tk.Frame(mapping_frame, bg="#f0f0f0")
        mapping_info_frame.pack(fill="x")
        
        self.mapping_label = tk.Label(
            mapping_info_frame,
            text=self.get_mapping_status_text(),
            font=("Arial", 9),
            bg="#f0f0f0",
            fg="#555",
            justify="left"
        )
        self.mapping_label.pack(side="left", padx=(0, 10))
        
        mapping_btn = tk.Button(
            mapping_info_frame,
            text="Configure Fields",
            command=self.configure_field_mapping,
            font=("Arial", 10),
            bg="#9b59b6",
            fg="white",
            cursor="hand2",
            padx=15,
            pady=5
        )
        mapping_btn.pack(side="left")
        
        # Excel file selection
        excel_frame = tk.LabelFrame(
            content_frame,
            text="Select Excel File",
            font=("Arial", 12, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50",
            padx=15,
            pady=15
        )
        excel_frame.pack(fill="x", pady=(0, 20))
        
        excel_entry = tk.Entry(
            excel_frame,
            textvariable=self.excel_path,
            font=("Arial", 10),
            width=50
        )
        excel_entry.pack(side="left", padx=(0, 10))
        
        excel_btn = tk.Button(
            excel_frame,
            text="Browse File",
            command=self.browse_excel,
            font=("Arial", 10),
            bg="#3498db",
            fg="white",
            cursor="hand2",
            padx=15,
            pady=5
        )
        excel_btn.pack(side="left")
        
        # Sheet selection
        sheet_frame = tk.LabelFrame(
            content_frame,
            text="Select Excel Sheet",
            font=("Arial", 12, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50",
            padx=15,
            pady=15
        )
        sheet_frame.pack(fill="x", pady=(0, 20))
        
        self.sheet_combo = ttk.Combobox(
            sheet_frame,
            textvariable=self.sheet_name,
            font=("Arial", 10),
            width=47,
            state="readonly"
        )
        self.sheet_combo.pack(side="left", padx=(0, 10))
        self.sheet_combo.set("Select a sheet or create new...")
        
        refresh_btn = tk.Button(
            sheet_frame,
            text="Refresh Sheets",
            command=self.load_sheets,
            font=("Arial", 10),
            bg="#9b59b6",
            fg="white",
            cursor="hand2",
            padx=15,
            pady=5
        )
        refresh_btn.pack(side="left", padx=(0, 10))
        
        clear_btn = tk.Button(
            sheet_frame,
            text="Clear Sheet Data",
            command=self.clear_sheet_data,
            font=("Arial", 10),
            bg="#e74c3c",
            fg="white",
            cursor="hand2",
            padx=15,
            pady=5
        )
        clear_btn.pack(side="left")
        
        # Progress frame
        self.progress_frame = tk.LabelFrame(
            content_frame,
            text="Progress",
            font=("Arial", 12, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50",
            padx=15,
            pady=15
        )
        self.progress_frame.pack(fill="both", expand=True, pady=(0, 20))
        
        self.progress_text = tk.Text(
            self.progress_frame,
            height=10,
            font=("Consolas", 9),
            bg="white",
            fg="#2c3e50",
            state="disabled",
            wrap="word"
        )
        self.progress_text.pack(fill="both", expand=True)
        
        # Scrollbar for progress text
        scrollbar = tk.Scrollbar(self.progress_text)
        scrollbar.pack(side="right", fill="y")
        self.progress_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.progress_text.yview)
        
        # Progress bar
        self.progress_bar = ttk.Progressbar(
            self.progress_frame,
            mode='indeterminate',
            length=300
        )
        
        # Convert button
        self.convert_btn = tk.Button(
            content_frame,
            text="Convert PDFs to Excel",
            command=self.start_conversion,
            font=("Arial", 14, "bold"),
            bg="#27ae60",
            fg="white",
            cursor="hand2",
            padx=30,
            pady=10
        )
        self.convert_btn.pack()
        
    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select Folder with PDF Files")
        if folder:
            self.folder_path.set(folder)
            
    def browse_excel(self):
        file = filedialog.asksaveasfilename(
            title="Select Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file:
            self.excel_path.set(file)
            self.load_sheets()
            
    def load_sheets(self):
        excel_path = self.excel_path.get()
        if not excel_path:
            messagebox.showwarning("Warning", "Please select an Excel file first!")
            return
            
        self.available_sheets = ["[Create New Sheet]"]
        
        if os.path.exists(excel_path):
            try:
                wb = openpyxl.load_workbook(excel_path)
                self.available_sheets.extend(wb.sheetnames)
                wb.close()
                self.sheet_combo['values'] = self.available_sheets
                self.sheet_combo.current(0)
            except Exception as e:
                messagebox.showerror("Error", f"Could not read Excel file:\n{str(e)}")
        else:
            self.sheet_combo['values'] = self.available_sheets
            self.sheet_combo.current(0)
            
    def clear_sheet_data(self):
        """Clear all data from the selected sheet (keeps headers)"""
        excel_path = self.excel_path.get()
        sheet_name = self.sheet_name.get()
        
        if not excel_path:
            messagebox.showwarning("Warning", "Please select an Excel file first!")
            return
        
        if not sheet_name or sheet_name == "Select a sheet or create new..." or sheet_name == "[Create New Sheet]":
            messagebox.showwarning("Warning", "Please select a sheet to clear!")
            return
        
        if not os.path.exists(excel_path):
            messagebox.showwarning("Warning", "Excel file does not exist yet!")
            return
        
        # Confirm action
        confirm = messagebox.askyesno(
            "Confirm Clear",
            f"Are you sure you want to clear all data from sheet '{sheet_name}'?\n\nThis will keep the headers but remove all entries."
        )
        
        if not confirm:
            return
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            if sheet_name not in wb.sheetnames:
                messagebox.showerror("Error", f"Sheet '{sheet_name}' not found!")
                wb.close()
                return
            
            ws = wb[sheet_name]
            
            # Delete all rows except the header (row 1)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
            
            wb.save(excel_path)
            wb.close()
            
            messagebox.showinfo("Success", f"Sheet '{sheet_name}' has been cleared!")
            self.log_message(f"\n✓ Cleared all data from sheet '{sheet_name}'")
            
        except PermissionError:
            messagebox.showerror("Error", "Cannot modify Excel file. Please close it in Excel and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
            
    def log_message(self, message):

        self.progress_text.config(state="normal")
        self.progress_text.insert("end", message + "\n")
        self.progress_text.see("end")
        self.progress_text.config(state="disabled")
        self.update_idletasks()
        
    def start_conversion(self):
        folder = self.folder_path.get()
        excel = self.excel_path.get()
        sheet = self.sheet_name.get()
        
        if not folder:
            messagebox.showerror("Error", "Please select a PDF folder!")
            return
            
        if not excel:
            messagebox.showerror("Error", "Please select an Excel file!")
            return
            
        if not sheet or sheet == "Select a sheet or create new...":
            messagebox.showerror("Error", "Please select a sheet!")
            return
            
        # Clear previous progress
        self.progress_text.config(state="normal")
        self.progress_text.delete(1.0, "end")
        self.progress_text.config(state="disabled")
        
        # Disable button and show progress bar
        self.convert_btn.config(state="disabled", text="Processing...")
        self.progress_bar.pack(pady=(10, 0))
        self.progress_bar.start()
        
        # Run conversion in a separate thread
        thread = threading.Thread(target=self.run_conversion, args=(folder, excel, sheet))
        thread.daemon = True
        thread.start()
        
    def run_conversion(self, folder_path, excel_path, sheet_name):
        try:
            # Ensure Excel file has .xlsx extension
            if not excel_path.lower().endswith('.xlsx'):
                excel_path += '.xlsx'
            
            # Get PDF files
            self.log_message(f"Searching for PDF files in: {folder_path}")
            pdf_files = get_pdf_files(folder_path)
            
            if not pdf_files:
                self.after(0, lambda: messagebox.showwarning("Warning", "No PDF files found!"))
                self.finish_conversion()
                return
            
            self.log_message(f"Found {len(pdf_files)} PDF file(s)\n")
            
            # Check if field mapping is configured
            if not self.field_mapping:
                self.log_message("⚠ No field mapping configured, using default (Total Amount)")
                use_default = True
            else:
                self.log_message(f"Using field mapping: {', '.join(self.field_mapping)}")
                use_default = False
            
            self.log_message("Extracting data from PDFs...")
            self.log_message("-" * 50)
            
            # Extract data from each PDF
            pdf_data = []
            for i, pdf_path in enumerate(pdf_files, 1):
                filename = os.path.basename(pdf_path)
                self.log_message(f"{i}. Processing: {filename}...")
                
                if use_default:
                    # Use old method for backward compatibility
                    total_amount = extract_total_amount(pdf_path)
                    pdf_data.append((filename, [total_amount], pdf_path))
                    self.log_message(f"   Total: {total_amount}")
                else:
                    # Use field mapping
                    field_values = extract_field_from_pdf(pdf_path, self.field_mapping)
                    
                    # Create ordered list of values matching field mapping
                    values = []
                    for field in self.field_mapping:
                        value = field_values.get(field, 'N/A')
                        values.append(value)
                        self.log_message(f"   {field}: {value}")
                    
                    pdf_data.append((filename, values, pdf_path))
            
            self.log_message("-" * 50)
            self.log_message(f"\nWriting to Excel file: {excel_path}")
            self.log_message(f"Sheet: {sheet_name}")
            
            # Write to Excel
            if use_default:
                # Use old write method
                old_format_data = [(name, vals[0], path) for name, vals, path in pdf_data]
                success = write_to_excel_gui(old_format_data, excel_path, sheet_name, self.log_message)
            else:
                # Use new write method with field mapping
                success = write_to_excel_with_mapping(
                    pdf_data, excel_path, sheet_name, self.field_mapping, self.log_message
                )
            
            if success:
                self.log_message("\n✓ Operation completed successfully!")
                self.after(0, lambda: messagebox.showinfo("Success", f"Successfully processed {len(pdf_data)} PDF files!\n\nExcel file saved at:\n{excel_path}\nSheet: {sheet_name}"))
            else:
                self.log_message("\n⚠ Operation was not completed.")
                
        except Exception as e:
            self.log_message(f"\n❌ Error: {str(e)}")
            self.after(0, lambda: messagebox.showerror("Error", f"An error occurred:\n{str(e)}"))
        finally:
            self.finish_conversion()
            
    def finish_conversion(self):
        self.progress_bar.stop()
        self.progress_bar.pack_forget()
        self.convert_btn.config(state="normal", text="Convert PDFs to Excel")
    
    def load_field_mapping(self):
        """Load saved field mapping from JSON file"""
        try:
            if os.path.exists(self.mapping_file):
                with open(self.mapping_file, 'r') as f:
                    data = json.load(f)
                    self.field_mapping = data.get('fields', [])
        except Exception as e:
            print(f"Could not load field mapping: {e}")
            self.field_mapping = []
    
    def save_field_mapping(self):
        """Save field mapping to JSON file"""
        try:
            with open(self.mapping_file, 'w') as f:
                json.dump({'fields': self.field_mapping}, f, indent=2)
        except Exception as e:
            print(f"Could not save field mapping: {e}")
    
    def get_mapping_status_text(self):
        """Get status text for field mapping"""
        if not self.field_mapping:
            return "⚠ No fields configured (will use default: Total Amount)"
        return f"✓ {len(self.field_mapping)} field(s) configured: {', '.join(self.field_mapping[:3])}{'...' if len(self.field_mapping) > 3 else ''}"
    
    def configure_field_mapping(self):
        """Open field mapping configuration dialog"""
        folder = self.folder_path.get()
        
        if not folder:
            messagebox.showwarning("Warning", "Please select a PDF folder first!")
            return
        
        # Get sample PDF
        pdf_files = get_pdf_files(folder)
        if not pdf_files:
            messagebox.showerror("Error", "No PDF files found in the selected folder!")
            return
        
        # Extract fields from first PDF as sample
        sample_pdf = pdf_files[0]
        self.log_message(f"Analyzing sample PDF: {os.path.basename(sample_pdf)}...")
        
        sample_fields = extract_all_fields_from_pdf(sample_pdf)
        
        if not sample_fields:
            messagebox.showwarning("Warning", "Could not extract fields from sample PDF!")
            return
        
        self.log_message(f"Found {len(sample_fields)} fields in sample PDF")
        
        # Open mapping dialog
        dialog = FieldMappingDialog(self, sample_fields, self.field_mapping)
        self.wait_window(dialog)
        
        # Get result
        result = dialog.get_result()
        if result is not None:
            self.field_mapping = result
            self.save_field_mapping()
            self.mapping_label.config(text=self.get_mapping_status_text())
            self.log_message(f"\n✓ Field mapping saved: {len(self.field_mapping)} field(s)")
            messagebox.showinfo("Success", f"Field mapping configured with {len(self.field_mapping)} field(s)!")

def write_to_excel_gui(pdf_data, excel_path, sheet_name, log_func):
    """
    Write PDF filenames, total amounts, and hyperlinks to an Excel file (GUI version).
    """
    try:
        existing_files = set()
        create_new_sheet = (sheet_name == "[Create New Sheet]")
        
        if os.path.exists(excel_path):
            try:
                wb = openpyxl.load_workbook(excel_path)
                
                if create_new_sheet:
                    # Generate new sheet name
                    base_name = "PDF Files"
                    counter = 1
                    while base_name in wb.sheetnames:
                        base_name = f"PDF Files {counter}"
                        counter += 1
                    ws = wb.create_sheet(base_name)
                    sheet_name = base_name
                    log_func(f"Creating new sheet: {sheet_name}")
                else:
                    # Use existing sheet
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        # Collect existing filenames by iterating only through cells with values
                        # This completely ignores empty/deleted rows
                        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
                            cell_value = row[0].value
                            if cell_value and str(cell_value).strip():
                                existing_files.add(cell_value)
                        
                        # Debug: Log what we found
                        if existing_files:
                            log_func(f"Found {len(existing_files)} existing file(s) in sheet")
                        else:
                            log_func("Sheet is empty, will add all files")
                    else:
                        ws = wb.create_sheet(sheet_name)
                
            except PermissionError:
                log_func(f"\n❌ ERROR: Cannot open '{excel_path}'")
                log_func("   The file is currently open in another program.")
                log_func("   Please close the file and try again.")
                return False
        else:
            wb = Workbook()
            ws = wb.active
            if create_new_sheet or not sheet_name:
                ws.title = "PDF Files"
                sheet_name = "PDF Files"
            else:
                ws.title = sheet_name
        
        # Add headers if new sheet or empty
        if ws.max_row == 1 or ws['A1'].value != "PDF Filename":
            ws['A1'] = "PDF Filename"
            ws['B1'] = "Total Amount"
            ws['C1'] = "Path to Invoice"
            ws['A1'].font = openpyxl.styles.Font(bold=True)
            ws['B1'].font = openpyxl.styles.Font(bold=True)
            ws['C1'].font = openpyxl.styles.Font(bold=True)
        
        new_data = [(name, amount, path) for name, amount, path in pdf_data if name not in existing_files]
        duplicates_count = len(pdf_data) - len(new_data)
        
        if duplicates_count > 0:
            log_func(f"\n⚠ Skipped {duplicates_count} duplicate file(s)")
        
        if not new_data:
            log_func("\n⚠ No new files to add (all files already exist in Excel)")
            return True
        
        start_row = ws.max_row + 1 if ws.max_row > 1 else 2
        
        for idx, (pdf_name, total_amount, pdf_path) in enumerate(new_data, start=start_row):
            ws[f'A{idx}'] = pdf_name
            ws[f'B{idx}'] = total_amount
            # Use relative path from Excel file location
            relative_path = os.path.relpath(pdf_path, os.path.dirname(excel_path))
            ws[f'C{idx}'].hyperlink = relative_path
            ws[f'C{idx}'].value = "Open Invoice"
            ws[f'C{idx}'].font = openpyxl.styles.Font(color="0563C1", underline="single")
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
        try:
            wb.save(excel_path)
            log_func(f"\n✓ Successfully wrote {len(new_data)} PDF file(s) to Excel")
            if duplicates_count > 0:
                log_func(f"  ({duplicates_count} duplicate(s) skipped)")
            return True
        except PermissionError:
            log_func(f"\n❌ ERROR: Cannot save to '{excel_path}'")
            log_func("   The file is currently open in another program.")
            return False
            
    except Exception as e:
        log_func(f"\n❌ Unexpected error: {e}")
        return False

def write_to_excel_with_mapping(pdf_data, excel_path, sheet_name, field_mapping, log_func):
    """
    Write PDF data to Excel using custom field mapping.
    
    Args:
        pdf_data: List of tuples (filename, [field_values], full_path)
        excel_path: Path where the Excel file will be saved
        sheet_name: Name of the sheet to write to
        field_mapping: List of field names (column headers)
        log_func: Function to log messages
    """
    try:
        existing_files = set()
        create_new_sheet = (sheet_name == "[Create New Sheet]")
        
        if os.path.exists(excel_path):
            try:
                wb = openpyxl.load_workbook(excel_path)
                
                if create_new_sheet:
                    # Generate new sheet name
                    base_name = "PDF Files"
                    counter = 1
                    while base_name in wb.sheetnames:
                        base_name = f"PDF Files {counter}"
                        counter += 1
                    ws = wb.create_sheet(base_name)
                    sheet_name = base_name
                    log_func(f"Creating new sheet: {sheet_name}")
                else:
                    # Use existing sheet
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        # Collect existing filenames
                        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
                            cell_value = row[0].value
                            if cell_value and str(cell_value).strip():
                                existing_files.add(cell_value)
                        
                        if existing_files:
                            log_func(f"Found {len(existing_files)} existing file(s) in sheet")
                        else:
                            log_func("Sheet is empty, will add all files")
                    else:
                        ws = wb.create_sheet(sheet_name)
                
            except PermissionError:
                log_func(f"\n❌ ERROR: Cannot open '{excel_path}'")
                log_func("   The file is currently open in another program.")
                log_func("   Please close the file and try again.")
                return False
        else:
            wb = Workbook()
            ws = wb.active
            if create_new_sheet or not sheet_name:
                ws.title = "PDF Files"
                sheet_name = "PDF Files"
            else:
                ws.title = sheet_name
        
        # Add headers if new sheet or empty
        expected_cols = 2 + len(field_mapping)  # PDF Filename + fields + Path
        if ws.max_row == 1 or ws['A1'].value != "PDF Filename":
            # Set up headers
            ws['A1'] = "PDF Filename"
            ws['A1'].font = openpyxl.styles.Font(bold=True)
            
            # Add field headers
            for idx, field_name in enumerate(field_mapping, start=2):
                col_letter = openpyxl.utils.get_column_letter(idx)
                ws[f'{col_letter}1'] = field_name
                ws[f'{col_letter}1'].font = openpyxl.styles.Font(bold=True)
            
            # Add path column
            path_col = openpyxl.utils.get_column_letter(len(field_mapping) + 2)
            ws[f'{path_col}1'] = "Path to Invoice"
            ws[f'{path_col}1'].font = openpyxl.styles.Font(bold=True)
        
        # Filter duplicates
        new_data = [(name, values, path) for name, values, path in pdf_data if name not in existing_files]
        duplicates_count = len(pdf_data) - len(new_data)
        
        if duplicates_count > 0:
            log_func(f"\n⚠ Skipped {duplicates_count} duplicate file(s)")
        
        if not new_data:
            log_func("\n⚠ No new files to add (all files already exist in Excel)")
            return True
        
        start_row = ws.max_row + 1 if ws.max_row > 1 else 2
        
        # Write data
        for row_idx, (pdf_name, field_values, pdf_path) in enumerate(new_data, start=start_row):
            # PDF filename
            ws[f'A{row_idx}'] = pdf_name
            
            # Field values
            for col_idx, value in enumerate(field_values, start=2):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws[f'{col_letter}{row_idx}'] = value
            
            # Path hyperlink
            path_col = openpyxl.utils.get_column_letter(len(field_mapping) + 2)
            relative_path = os.path.relpath(pdf_path, os.path.dirname(excel_path))
            ws[f'{path_col}{row_idx}'].hyperlink = relative_path
            ws[f'{path_col}{row_idx}'].value = "Open Invoice"
            ws[f'{path_col}{row_idx}'].font = openpyxl.styles.Font(color="0563C1", underline="single")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        for col_idx in range(2, len(field_mapping) + 2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
        path_col = openpyxl.utils.get_column_letter(len(field_mapping) + 2)
        ws.column_dimensions[path_col].width = 20
        
        try:
            wb.save(excel_path)
            log_func(f"\n✓ Successfully wrote {len(new_data)} PDF file(s) to Excel")
            if duplicates_count > 0:
                log_func(f"  ({duplicates_count} duplicate(s) skipped)")
            return True
        except PermissionError:
            log_func(f"\n❌ ERROR: Cannot save to '{excel_path}'")
            log_func("   The file is currently open in another program.")
            return False
            
    except Exception as e:
        log_func(f"\n❌ Unexpected error: {e}")
        return False

if __name__ == "__main__":
    main()
